from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np

# Generate example data
data = np.random.rand(5, 5)
df = pd.DataFrame(data, columns=['number', 'value1', 'value2', 'value3', 'value4'])

# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Define starting positions for the data blocks
start_rows = [1, 9]  # Rows where data blocks start
start_cols = [1, 7, 13]  # Columns where data blocks start

# Define the color scale rule for the heatmap
color_scale_rule = ColorScaleRule(start_type='num', start_value=0, start_color='FFF8696B',
                                  mid_type='percentile', mid_value=50, mid_color='FFFFEB84',
                                  end_type='num', end_value=1, end_color='FF63BE7B')
# Define the color scale rule for the heatmap
color_scale_rule = ColorScaleRule(start_type='num', start_value=0, start_color='FF0000',  # Red for the lowest values
                                  mid_type='num', mid_value=0.5, mid_color='FFFF00',    # Yellow for the midpoint
                                  end_type='num', end_value=1, end_color='00FF00')      # Green for the highest values

# Write data and headers to the worksheet
for start_row in start_rows:
    for start_col in start_cols:
        # Write the headers
        for col_index, header in enumerate(df.columns, start=start_col):
            cell = ws.cell(row=start_row, column=col_index, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type="solid")
        
        # Write the data rows and apply the heatmap color scale
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row + 1):
            for c_idx, value in enumerate(row, start=start_col):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                # Apply the color scale rule to data cells only, not the header
                ws.conditional_formatting.add(cell.coordinate, color_scale_rule)

# Insert the descriptive text between the data blocks
for i, col in enumerate(start_cols[1:], start=1):
    ws.cell(row=start_rows[1] - 1, column=col, value='Next column table')
    ws.cell(row=start_rows[1] - 1, column=col).font = Font(bold=True)

# Save the workbook
file_path = 'complex_heatmap_openpyxl.xlsx'
wb.save(file_path)
